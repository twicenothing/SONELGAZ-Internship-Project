import pandas as pd
import numpy as np 
from region_funcs import *
from region_funcs_accroissement import * 
from region_funcs_apport import * 
from region_funcs_gaz import * 
from region_funcs_gaz_accroissement import * 
from region_funcs_gaz_apport import *
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from ventes.region_ventes import *
from ventes.region_ventes_gaz import * 
from RCN.rdc_TB import *
from creances.solde_df import *
from creances.prise_en_charge_df import *
from creances.encaissement_df import *
from creances.solde23_df import * 
from creances.recettes_df import *
from creances.ddc_df import ddc_tb
from creances.old_data import *
from elec.elec_tb import *
from data import TB


pd.set_option('display.float_format', '{:.2f}'.format)

def load_old_data_TB(file=TB):
    df = pd.read_excel(file,sheet_name='élec')
    df = df.dropna(how='all')
    first_row = df.columns.to_list()
    df.loc[0] = first_row  # Temporarily add the row with index -1
    def detect_tables(df, keywords):
        titles = []
        for index, row in df.iterrows():
            
            # Check if any cell in the row contains one of the keywords
            if row.astype(str).str.contains('|'.join(keywords), case=False).any():
                
                titles.append(index)
        
        # Return the list of detected titles after the loop finishes
        return titles
    keywords=['Nombre d\'abonnés']
    table_titles = detect_tables(df, keywords)
    table_titles
    #Extraction de la table nombre d'abonnés
    start = table_titles[0]+1
    end = table_titles[0]+14
    nombre_abbo2013 = df.iloc[start:end]
    nombre_abbo2013.columns = nombre_abbo2013.iloc[0]
    nombre_abbo2013 = nombre_abbo2013.iloc[1:]
    nombre_abbo2013.index = nombre_abbo2013['DD']
    del nombre_abbo2013['DD']
    nombre_abbo2013.index.name = None
    nombre_abbo2013.columns.naame = None
    nombre_abbo2013.columns = np.arange(0,58)
    nombre_abbo2013 = nombre_abbo2013.loc[:,:47]
    nombre_abbo2013.columns = nombre_abbo2013.iloc[0]
    nombre_abbo2013 = nombre_abbo2013[1:]
    nombre_abbo2013 = nombre_abbo2013.replace(0,np.nan)
    nombre_abbo2013 = nombre_abbo2013.dropna(axis=1, how='all' )
    french_months = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]
    # Number of columns per month
    columns_per_month = 4

    # Calculate the number of months based on the columns
    num_months = (len(nombre_abbo2013.columns) + columns_per_month - 1) // columns_per_month

    # Generate month labels dynamically up to the current number of months
    months = french_months[:num_months]

    # Create the high-level (month) and low-level (original column names) indices
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)

    # Truncate high-level index to match the number of columns
    high_level = high_level[:len(nombre_abbo2013.columns)]

    # Create a MultiIndex
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, nombre_abbo2013.columns)]

    )

    # Set the MultiIndex as the new columns
    nombre_abbo2013.columns = multi_index
    nombre_abbo2013.columns.name = None
    nombre_abbo2013.index.name = None

    keywords=['Nombre d\'abonnés','Accroissement']
    table_titles = detect_tables(df, keywords)
    start = table_titles[1]+1
    end = table_titles[1]+14
    accroissement2013 = df.iloc[start:end]
    accroissement2013 = accroissement2013.dropna(axis='columns',how='all')
    accroissement2013 = accroissement2013.iloc[:,:53]
    accroissement2013 = accroissement2013.replace(0,np.nan)
    accroissement2013.columns = accroissement2013.iloc[1]
    accroissement2013 = accroissement2013.iloc[2:]
    accroissement2013 = accroissement2013.dropna(axis='columns',how='all')

    accroissement2013.columns.values[0]= 'placeholder'
    accroissement2013.index = accroissement2013['placeholder']
    del accroissement2013['placeholder']

    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(accroissement2013.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month

    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]

    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)

    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)

    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(accroissement2013.columns)]

    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, accroissement2013.columns)]

    )

    # Set the MultiIndex as the new columns
    accroissement2013.columns = multi_index
    accroissement2013.index.name = None

        #Extraction de la table apport
    keywords=['Nombre d\'abonnés','Accroissement','Apport']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]+1
    end = table_titles[2]+13
    apport2023 = df.iloc[start:end]
    apport2023 = apport2023.dropna(axis='columns',how='all')
    apport2023 = apport2023.iloc[:,:53]
    apport2023 = apport2023.replace(0,np.nan)
    apport2023.columns = apport2023.iloc[0]
    apport2023 = apport2023.iloc[1:]
    apport2023.columns.values[0] = 'placeholder'
    apport2023.index = apport2023['placeholder']
    del apport2023['placeholder']
    apport2023 = apport2023.dropna(axis=1,how='all')
    apport2023.index.name = None
    apport2023.columns.name = None

    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(apport2023.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month

    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]

    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)

    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)

    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(apport2023.columns)]

    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, apport2023.columns)]

    )

    # Set the MultiIndex as the new columns
    apport2023.columns = multi_index
    apport2023.index.name = None
    apport2023 = apport2023.replace(np.nan, 0)


    keywords=['Nombre d\'abonnés','Accroissement','II-  Résiliation']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]+1
    end = table_titles[2]+13
    resil = df.iloc[start:end]
    resil = resil.dropna(axis='columns',how='all')
    resil = resil.iloc[:,:53]
    resil = resil.replace(0,np.nan)
    resil.columns = resil.iloc[0]
    resil = resil.iloc[1:]
    resil.columns.values[0] = 'placeholder'
    resil.index = resil['placeholder']
    del resil['placeholder']
    resil = resil.dropna(axis=1,how='all')
    resil.index.name = None
    resil.columns.name = None
    
    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(resil.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month
    
    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]
    
    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)
    
    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)
    
    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(resil.columns)]
    
    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
    [(high, col) for high, col in zip(high_level, resil.columns)]
    
    )
    
    # Set the MultiIndex as the new columns
    resil.columns = multi_index
    resil.index.name = None
    resil = resil.replace(np.nan, 0)


    keywords=['Nombre d\'abonnés','Accroissement','II-  Réabonnement']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]
    end = table_titles[2]+12
    reabonne = df.iloc[start:end]
    reabonne = reabonne.dropna(axis='columns',how='all')
    reabonne = reabonne.iloc[:,:53]
    reabonne = reabonne.replace(0,np.nan)
    reabonne.columns = reabonne.iloc[0]
    reabonne = reabonne.iloc[1:]
    reabonne.columns.values[0] = 'placeholder'
    reabonne.index = reabonne['placeholder']
    del reabonne['placeholder']
    reabonne = reabonne.dropna(axis=1, how='all')
    reabonne.index.name = None
    reabonne.columns.name = None
    
    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(reabonne.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month
    
    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]
    
    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)
    
    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)
    
    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(reabonne.columns)]
    
    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, reabonne.columns)]
    )
    
    # Set the MultiIndex as the new columns
    reabonne.columns = multi_index
    reabonne.index.name = None
    reabonne = reabonne.replace(np.nan, 0)

    keywords=['Nombre d\'abonnés','Accroissement','Apport nouveau']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]-1
    end = table_titles[2]+11
    apport_nouveau = df.iloc[start:end]
    apport_nouveau = apport_nouveau.dropna(axis='columns', how='all')
    apport_nouveau = apport_nouveau.iloc[:, :53]
    apport_nouveau = apport_nouveau.replace(0, np.nan)
    apport_nouveau.columns = apport_nouveau.iloc[0]
    apport_nouveau = apport_nouveau.iloc[1:]
    apport_nouveau.columns.values[0] = 'placeholder'
    apport_nouveau.index = apport_nouveau['placeholder']
    del apport_nouveau['placeholder']
    apport_nouveau = apport_nouveau.dropna(axis=1, how='all')
    apport_nouveau.index.name = None
    apport_nouveau.columns.name = None
    
    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(apport_nouveau.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month
    
    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]
    
    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)
    
    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)
    
    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(apport_nouveau.columns)]
    
    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
    [(high, col) for high, col in zip(high_level, apport_nouveau.columns)]
    )
    
    # Set the MultiIndex as the new columns
    apport_nouveau.columns = multi_index
    apport_nouveau.index.name = None
    apport_nouveau = apport_nouveau.replace(np.nan, 0)
    
    keywords=['Nombre d\'abonnés','Accroissement','III-  Ventes électricité']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]-3
    end = table_titles[2]+9
    ventes = df.iloc[start:end]
    ventes = ventes.dropna(axis='columns', how='all')
    ventes = ventes.iloc[:, :53]
    ventes = ventes.replace(0, np.nan)
    ventes.columns = ventes.iloc[0]
    ventes = ventes.iloc[1:]
    ventes.columns.values[0] = 'placeholder'
    ventes.index = ventes['placeholder']
    del ventes['placeholder']
    ventes = ventes.dropna(axis=1, how='all')
    ventes.index.name = None
    ventes.columns.name = None
    
    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(ventes.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month
    
    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]
    
    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)
    
    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)
    
    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(ventes.columns)]
    
    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, ventes.columns)]
    )
    
    # Set the MultiIndex as the new columns
    ventes.columns = multi_index
    ventes.index.name = None
    ventes = ventes.replace(np.nan, 0)

    keywords=['Nombre d\'abonnés','Accroissement','IV-  Achats électricité & pertes d\'énergie']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]-3
    end = table_titles[2]+9
    achats = df.iloc[start:end]
    achats = achats.dropna(axis='columns', how='all')
    achats = achats.iloc[:, :53]
    achats = achats.replace(0, np.nan)
    achats.columns = achats.iloc[0]
    achats = achats.iloc[1:]
    achats.columns.values[0] = 'placeholder'
    achats.index = achats['placeholder']
    del achats['placeholder']
    achats = achats.dropna(axis=1, how='all')
    achats.index.name = None
    achats.columns.name = None
    
    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(achats.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month
    
    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]
    
    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)
    
    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)
    
    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(achats.columns)]
    
    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, achats.columns)]
    )
    
    # Set the MultiIndex as the new columns
    achats.columns = multi_index
    achats.index.name = None
    achats = achats.replace(np.nan, 0)

    keywords=['Nombre d\'abonnés','Accroissement','V-  Chiffre d\'affaires électricité']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]-3
    end = table_titles[2]+9
    chiffre_aff = df.iloc[start:end]
    chiffre_aff = chiffre_aff.dropna(axis='columns', how='all')
    chiffre_aff = chiffre_aff.iloc[:, :53]
    chiffre_aff = chiffre_aff.replace(0, np.nan)
    chiffre_aff.columns = chiffre_aff.iloc[0]
    chiffre_aff = chiffre_aff.iloc[1:]
    chiffre_aff.columns.values[0] = 'placeholder'
    chiffre_aff.index = chiffre_aff['placeholder']
    del chiffre_aff['placeholder']
    chiffre_aff = chiffre_aff.dropna(axis=1, how='all')
    chiffre_aff.index.name = None
    chiffre_aff.columns.name = None
    
    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(chiffre_aff.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month
    
    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]
    
    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)
    
    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)
    
    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(chiffre_aff.columns)]
    
    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, chiffre_aff.columns)]
    )
    
    # Set the MultiIndex as the new columns
    chiffre_aff.columns = multi_index
    chiffre_aff.index.name = None
    chiffre_aff = chiffre_aff.replace(np.nan, 0)



    keywords=['Nombre d\'abonnés','Accroissement','VI- Prix de Ventes Moyen électricité']
    table_titles = detect_tables(df, keywords)
    start = table_titles[2]-3
    end = table_titles[2]+9
    prix = df.iloc[start:end]
    prix = prix.dropna(axis='columns', how='all')
    prix = prix.iloc[:, :53]
    prix = prix.replace(0, np.nan)
    prix.columns = prix.iloc[0]
    prix = prix.iloc[1:]
    prix.columns.values[0] = 'placeholder'
    prix.index = prix['placeholder']
    del prix['placeholder']
    prix = prix.dropna(axis=1, how='all')
    prix.index.name = None
    prix.columns.name = None
    
    # Separate the last 4 columns for CUMUL
    num_regular_columns = len(prix.columns) - 4  # Exclude the last 4 for now
    num_months = (num_regular_columns + columns_per_month - 1) // columns_per_month
    
    # Generate month labels dynamically in French up to the current number of months
    months = french_months[:num_months]
    
    # Create the high-level index for the regular columns
    high_level = []
    for month in months:
        high_level.extend([month] * columns_per_month)
    
    # Add 'CUMUL' for the last 4 columns
    high_level.extend(['CUMUL'] * 4)
    
    # Ensure the high-level index matches the number of columns
    high_level = high_level[:len(prix.columns)]
    
    # Combine high-level index with existing column names (low-level index)
    multi_index = pd.MultiIndex.from_tuples(
        [(high, col) for high, col in zip(high_level, prix.columns)]
    )
    
    # Set the MultiIndex as the new columns
    prix.columns = multi_index
    prix.index.name = None
    prix = prix.replace(np.nan, 0)

    return nombre_abbo2013,accroissement2013,apport2023,resil,reabonne,apport_nouveau,ventes,achats,chiffre_aff,prix

















def TB_clientele():
    w = region_nombre_abonne()
    nombre_abbo_gaz = region_nombre_abonne_gaz()

    # Save the first DataFrame to Excel
    file_name = 'TB_test.xlsx'
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        w.to_excel(writer, sheet_name='Clientele', startrow=4)

    # Customize the first DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name='Clientele',
        start_row=5,  # Data starts from row 5
        title="Nombre d'abonnés électricité",
        last_row_color="87CEEB"  # Light blue for the last row
    )

    # Add a gap of 4 rows before the second table
    gap_row_2nd = w.shape[0] + 9  # Adds a 4-row gap between the two tables

    # Save the second DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        nombre_abbo_gaz.to_excel(writer, sheet_name='Clientele', startrow=gap_row_2nd, startcol=0)

    # Customize the second DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name='Clientele',
        start_row=gap_row_2nd + 1,  # Adjust the row number after the second DataFrame
        title="Nombre d'abonnés gaz",
        last_row_color="90EE90"  # Light green for the last row
    )

    # Add a gap of 4 rows before the third table
    gap_row_3rd = gap_row_2nd + nombre_abbo_gaz.shape[0] + 5  # Adds a 4-row gap between the second and third table

    accroissement_elec = region_accroissement()
    # Save the third DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_elec.to_excel(writer, sheet_name='Clientele', startrow=gap_row_3rd, startcol=0)

    # Customize the third DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name='Clientele',
        start_row=gap_row_3rd + 1,  # Adjust the row number after the third DataFrame
        title="Accroissement abonnées électricité",
        last_row_color="FFC0CB"  # Light pink for the last row (change this color as needed)
    )

    # Add a gap of 4 rows before the fourth table
    gap_row_4rth = gap_row_3rd + accroissement_elec.shape[0] + 5  # Adds a 4-row gap between the third and fourth table

    accroissement_gaz = region_accroissement_gaz()
    # Save the fourth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_gaz.to_excel(writer, sheet_name='Clientele', startrow=gap_row_4rth, startcol=0)

    # Customize the fourth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name='Clientele',
        start_row=gap_row_4rth + 1,  # Adjust the row number after the fourth DataFrame
        title="Accroissement abonnées gaz",
        last_row_color="D8B7DD"  # Light pink for the last row (change this color as needed)
    )

    # Add a gap of 4 rows before the fifth table
    gap_row_5th = gap_row_4rth + accroissement_gaz.shape[0] + 5  # Adds a 4-row gap between the fourth and fifth table

    apport_elec = region_apport()
    # Save the fifth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        apport_elec.to_excel(writer, sheet_name='Clientele', startrow=gap_row_5th, startcol=0)

    # Customize the fifth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name='Clientele',
        start_row=gap_row_5th + 1,  # Adjust the row number after the fifth DataFrame
        title="Nouveaux abonnées électricité",
        last_row_color="FFFFE0"  # Light yellow for the last row
    )
    gap_row_6th = gap_row_5th + apport_elec.shape[0] + 5  # Adds a 4-row gap between the fourth and fifth table

    apport_gaz = region_apport_gaz()
    # Save the fifth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        apport_gaz.to_excel(writer, sheet_name='Clientele', startrow=gap_row_6th, startcol=0)

    # Customize the fifth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name='Clientele',
        start_row=gap_row_6th + 1,  # Adjust the row number after the fifth DataFrame
        title="Nouveaux abonnées gaz",
        last_row_color="FFA500"  # Light yellow for the last row
    )



def customize_excel_table(file_name, sheet_name, start_row, title, last_row_color):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    last_row_fill = PatternFill(start_color=last_row_color, end_color=last_row_color, fill_type='solid')

    # Add a title
    ws.merge_cells(start_row=start_row - 1, start_column=1, end_row=start_row - 1, end_column=ws.max_column)
    title_cell = ws.cell(row=start_row - 1, column=1)
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = centered_alignment

    # Adjust column width and row height for the table
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 15  # Adjust column width

    for row in range(start_row, ws.max_row + 1):
        ws.row_dimensions[row].height = 20  # Adjust row height

    # Apply borders and style to the table
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Color the last row
    for cell in ws[ws.max_row]:
        cell.fill = last_row_fill

    # Save the workbook
    wb.save(file_name)



def TB_ventes():
    # Adding a new sheet for the 7th DataFrame
        new_sheet_name = 'ventes_achats_élec'  # The name for the new sheet (can be customized)
        gap_row_7th = 4  # Start at a custom row number for the new sheet

        # Fetch the 7th DataFrame (replace with your actual data function)
        new_sheet_data = region_ventes()  # Replace with the function for your 7th DataFrame
        with pd.ExcelWriter('TB_test.xlsx', engine='openpyxl', mode='a') as writer:
            # Write the new DataFrame to the new sheet
            new_sheet_data.to_excel(writer, sheet_name=new_sheet_name, startrow=gap_row_7th, startcol=0)

        # Customize the new DataFrame in the new sheet
        customize_excel_table(
            file_name='TB_test.xlsx',
            sheet_name=new_sheet_name,
            start_row=gap_row_7th + 1,  # Adjust the row number after the 7th DataFrame
            title="Ventes & Achats électricité (BT/MT en GWh) du mois",  # Title for the new DataFrame
            last_row_color="C8A2D4"  # Light yellow for the last row
        )
        # Assuming the 7th DataFrame was saved to the new sheet 'OtherSheet' previously
        gap_row_8th = gap_row_7th + new_sheet_data.shape[0] + 5  # Adds a 4-row gap after the 7th DataFrame
        ventes_cumul = get_cumul_ventes()
        # Save the 8th DataFrame to the new sheet, right below the 7th DataFrame
        with pd.ExcelWriter('TB_test.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            ventes_cumul.to_excel(writer, sheet_name=new_sheet_name, startrow=gap_row_8th, startcol=0)

        # Customize the 8th DataFrame in the Excel file
        customize_excel_table(
            file_name='TB_test.xlsx',
            sheet_name=new_sheet_name,
            start_row=gap_row_8th + 1,  # Adjust the row number after the 8th DataFrame
            title="Ventes & Achats électricité (BT/MT en GWh) cumulés",  # Update this title accordingly
            last_row_color="C8A2D4"  # Yellow gold, or change to any color you prefer
        )

def TB_ventes_gaz():
    # Adding a new sheet for the 7th DataFrame
        new_sheet_name = 'ventes_achats_gaz'  # The name for the new sheet (can be customized)
        gap_row_7th = 4  # Start at a custom row number for the new sheet

        # Fetch the 7th DataFrame (replace with your actual data function)
        new_sheet_data = region_ventes_gaz()  # Replace with the function for your 7th DataFrame
        with pd.ExcelWriter('TB_test.xlsx', engine='openpyxl', mode='a') as writer:
            # Write the new DataFrame to the new sheet
            new_sheet_data.to_excel(writer, sheet_name=new_sheet_name, startrow=gap_row_7th, startcol=0)

        # Customize the new DataFrame in the new sheet
        customize_excel_table(
            file_name='TB_test.xlsx',
            sheet_name=new_sheet_name,
            start_row=gap_row_7th + 1,  # Adjust the row number after the 7th DataFrame
            title="Ventes & Achats GAZ (BP/MP en Mth) du mois",  # Title for the new DataFrame
            last_row_color="C8A2D4"  # Light yellow for the last row
        )
        # Assuming the 7th DataFrame was saved to the new sheet 'OtherSheet' previously
        gap_row_8th = gap_row_7th + new_sheet_data.shape[0] + 5  # Adds a 4-row gap after the 7th DataFrame
        ventes_cumul = get_cumul_ventes_gaz()
        # Save the 8th DataFrame to the new sheet, right below the 7th DataFrame
        with pd.ExcelWriter('TB_test.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            ventes_cumul.to_excel(writer, sheet_name=new_sheet_name, startrow=gap_row_8th, startcol=0)

        # Customize the 8th DataFrame in the Excel file
        customize_excel_table(
            file_name='TB_test.xlsx',
            sheet_name=new_sheet_name,
            start_row=gap_row_8th + 1,  # Adjust the row number after the 8th DataFrame
            title="Ventes  & Achats GAZ (BP/ MP en Mth) cumulés",  # Update this title accordingly
            last_row_color="C8A2D4"  # Yellow gold, or change to any color you prefer
        )

def TB_RCN_elec():
    w = extension_portfeuille_elec()
    nombre_abbo_gaz = branchements_simples_elecrticite()
    name='RCN_Elec'
    # Save the first DataFrame to Excel
    file_name = 'TB_test.xlsx'
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        w.to_excel(writer, sheet_name=name, startrow=4)

    # Customize the first DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=5,  # Data starts from row 5
        title="Extension Portfeuille",
        last_row_color="87CEEB"  # Light blue for the last row
    )

    # Add a gap of 4 rows before the second table
    gap_row_2nd = w.shape[0] + 9  # Adds a 4-row gap between the two tables

    # Save the second DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        nombre_abbo_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_2nd, startcol=0)

    # Customize the second DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_2nd + 1,  # Adjust the row number after the second DataFrame
        title="Branchements Simples Electricité",
        last_row_color="90EE90"  # Light green for the last row
    )
     # Add a gap of 4 rows before the third table
    gap_row_3rd = gap_row_2nd + nombre_abbo_gaz.shape[0] + 5  # Adds a 4-row gap between the second and third table

    accroissement_elec = extension_affaires_elec()
    # Save the third DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_elec.to_excel(writer, sheet_name=name, startrow=gap_row_3rd, startcol=0)

    # Customize the third DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_3rd + 1,  # Adjust the row number after the third DataFrame
        title="Extension Affaires",
        last_row_color="FFC0CB"  # Light pink for the last row (change this color as needed)
    )
    # Add a gap of 4 rows before the fourth table
    gap_row_4rth = gap_row_3rd + accroissement_elec.shape[0] + 5  # Adds a 4-row gap between the third and fourth table

    accroissement_gaz = branchment_affaires_elec()
    # Save the fourth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_4rth, startcol=0)

    # Customize the fourth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_4rth + 1,  # Adjust the row number after the fourth DataFrame
        title="Branchement affaires",
        last_row_color="D8B7DD"  # Light pink for the last row (change this color as needed)
    )


def TB_RCN_gaz():
    w = extension_portfeuille_gaz()
    nombre_abbo_gaz = branchements_simples_gaz()
    name='RCN_Gaz'
    # Save the first DataFrame to Excel
    file_name = 'TB_test.xlsx'
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        w.to_excel(writer, sheet_name=name, startrow=4)

    # Customize the first DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=5,  # Data starts from row 5
        title="Extension Portfeuille",
        last_row_color="87CEEB"  # Light blue for the last row
    )

    # Add a gap of 4 rows before the second table
    gap_row_2nd = w.shape[0] + 9  # Adds a 4-row gap between the two tables

    # Save the second DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        nombre_abbo_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_2nd, startcol=0)

    # Customize the second DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_2nd + 1,  # Adjust the row number after the second DataFrame
        title="Branchements Simples Gaz",
        last_row_color="90EE90"  # Light green for the last row
    )
 # Add a gap of 4 rows before the third table
    gap_row_3rd = gap_row_2nd + nombre_abbo_gaz.shape[0] + 5  # Adds a 4-row gap between the second and third table

    accroissement_elec = extension_affaires_gaz()
    # Save the third DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_elec.to_excel(writer, sheet_name=name, startrow=gap_row_3rd, startcol=0)

    # Customize the third DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_3rd + 1,  # Adjust the row number after the third DataFrame
        title="Extension Affaires",
        last_row_color="FFC0CB"  # Light pink for the last row (change this color as needed)
    )
    # Add a gap of 4 rows before the fourth table
    gap_row_4rth = gap_row_3rd + accroissement_elec.shape[0] + 5  # Adds a 4-row gap between the third and fourth table

    accroissement_gaz = branchment_affaires_gaz()
    # Save the fourth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_4rth, startcol=0)

    # Customize the fourth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_4rth + 1,  # Adjust the row number after the fourth DataFrame
        title="Branchement affaires",
        last_row_color="D8B7DD"  # Light pink for the last row (change this color as needed)
    )


def TB_solde():
    w = solde_dataframe()
    name='calcul-DCC-TX-Encais'
    # Save the first DataFrame to Excel
    file_name = 'TB_test.xlsx'
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        w.to_excel(writer, sheet_name=name, startrow=4)

    # Customize the first DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=5,  # Data starts from row 5
        title="2024",
        last_row_color="87CEEB"  # Light blue for the last row
    )
     # Add a gap of 4 rows before the second table
    gap_row_2nd = w.shape[0] + 9  # Adds a 4-row gap between the two tables
    nombre_abbo_gaz = prise_en_charge_dataframe()
    # Save the second DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        nombre_abbo_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_2nd, startcol=0)

    # Customize the second DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_2nd + 1,  # Adjust the row number after the second DataFrame
        title="2024",
        last_row_color="90EE90"  # Light green for the last row
    )
    gap_row_3rd = gap_row_2nd + nombre_abbo_gaz.shape[0] + 5  # Adds a 4-row gap between the second and third table

    accroissement_elec = encaissement_dataframe()
    # Save the third DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_elec.to_excel(writer, sheet_name=name, startrow=gap_row_3rd, startcol=0)

    # Customize the third DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_3rd + 1,  # Adjust the row number after the third DataFrame
        title="2024",
        last_row_color="FFC0CB"  # Light pink for the last row (change this color as needed)
    )

  

    solde23= load_old_data_creance()
    solde23.columns = w.columns
    solde23 = solde23.droplevel([0, 1], axis=1)
    accroissement_elec = accroissement_elec.droplevel([0, 1], axis=1)
    nombre_abbo_gaz = nombre_abbo_gaz.droplevel([0, 1], axis=1)
    nombre_abbo_gaz.replace(0,np.nan,inplace=True)
    solde23.replace(0,np.nan,inplace=True)


    tx_encaissement = (accroissement_elec / (solde23 + nombre_abbo_gaz))* 100
    tx_encaissement = tx_encaissement.fillna(0)

    # Correct calculation for gap_row_4rth
    gap_row_4rth = gap_row_3rd + accroissement_elec.shape[0] + 5  # Adds a 4-row gap after the third DataFrame

    # Save the fourth DataFrame (tx_encaissement) to the same Excel sheet
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        tx_encaissement.to_excel(writer, sheet_name=name, startrow=gap_row_4rth, startcol=0)

    # Customize the fourth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_4rth + 1,  # Adjust the row number for the title
        title="2024",
        last_row_color="D8B7DD"  # Light purple for the last row
    )
    # Add a gap of 4 rows before the fifth table
    gap_row_5th = gap_row_4rth + tx_encaissement.shape[0] + 5  # Adds a 4-row gap between the fourth and fifth table

    apport_elec = recettes_tb()
    # Save the fifth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        apport_elec.to_excel(writer, sheet_name=name, startrow=gap_row_5th, startcol=0)

    # Customize the fifth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_5th + 1,  # Adjust the row number after the fifth DataFrame
        title="2024",
        last_row_color="FFFFE0"  # Light yellow for the last row
    )
    gap_row_6th = gap_row_5th + apport_elec.shape[0] + 5  # Adds a 4-row gap between the fourth and fifth table

    apport_gaz = ddc_tb()
    # Save the fifth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        apport_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_6th, startcol=0)

    # Customize the fifth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_6th + 1,  # Adjust the row number after the fifth DataFrame
        title="2024 DDC",
        last_row_color="FFA500"  # Light yellow for the last row
    )
    solde23 = load_solde_23()
    solde23.columns = w.columns
    w.replace(0,np.nan,inplace=True)
    solde23.replace(0,np.nan,inplace=True)
    evolution = (w - solde23)/solde23*100

    gap_row_7th = gap_row_6th + apport_gaz.shape[0] + 5  # Adds a 4-row gap between the fourth and fifth table

    # Save the fifth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        evolution.to_excel(writer, sheet_name=name, startrow=gap_row_7th, startcol=0)

    # Customize the fifth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_7th + 1,  # Adjust the row number after the fifth DataFrame
        title="Evolution solde /mois",
        last_row_color="00FFFF"  # Light yellow for the last row
    )

    solde_old = load_old_data_creance()
    solde_old.columns = w.columns
    w.replace(0,np.nan,inplace=True)
    solde_old.replace(0,np.nan,inplace=True)
    evolution2 = (w - solde_old)/solde_old*100
    gap_row_8th = gap_row_7th + evolution.shape[0] + 5  # Adds a 4-row gap between the fourth and fifth table

    # Save the fifth DataFrame to the same Excel sheet, with a gap between tables row-wise
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        evolution2.to_excel(writer, sheet_name=name, startrow=gap_row_8th, startcol=0)

    # Customize the fifth DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_8th + 1,  # Adjust the row number after the fifth DataFrame
        title="Evolution solde /fin d'année",
        last_row_color="D3D3D3"  # Light yellow for the last row
    )



def TB_elec():
   # Load data for tables
    n, a, ap, r, rea, apnv, ventesTB, achatsTB, chiffaf, pr = load_old_data_TB()
    w = nombre_abo_tb()
    nombre_abbo_gaz = accroissement_tb()
    accroissement_elec = resiliation_tb()
    tx_encaissement = apport_tb()
    apport_elec = reabonne_tb()
    apport_gaz = apport_nv_tb()
    evolution = ventes_tb()
    evolution2 = achats_tb()
    chiffre_aff = chiffre_aff_tb()
    prix = prix_tb()

    # Sheet name and file name
    name = 'élec'
    file_name = 'TB_test.xlsx'

    # Save the first DataFrame (n) to Excel
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        n.to_excel(writer, sheet_name=name, startrow=4)

    # Customize the first DataFrame in the Excel file
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=5,
        title="2024",
        last_row_color="87CEEB"
    )

    # Add a gap of 4 rows before the second table
    gap_row_2nd = n.shape[0] + 9
    # Save second DataFrame (w)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        w.to_excel(writer, sheet_name=name, startrow=gap_row_2nd, startcol=0)

    # Customize second DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_2nd + 1,
        title="2024",
        last_row_color="90EE90"
    )

    # Add a gap of 4 rows before the third table
    gap_row_3rd = gap_row_2nd + w.shape[0] + 5
    # Save third DataFrame (a)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        a.to_excel(writer, sheet_name=name, startrow=gap_row_3rd, startcol=0)

    # Customize third DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_3rd + 1,
        title="2024",
        last_row_color="FFC0CB"
    )

    # Add a gap of 4 rows before the fourth table
    gap_row_4th = gap_row_3rd + a.shape[0] + 5
    # Save fourth DataFrame (nombre_abbo_gaz)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        nombre_abbo_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_4th, startcol=0)

    # Customize fourth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_4th + 1,
        title="2024",
        last_row_color="FFFFE0"
    )



    # Add a gap of 4 rows before the fifth table
    gap_row_5th = gap_row_4th + nombre_abbo_gaz.shape[0] + 5
    # Save fifth DataFrame (r)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        r.to_excel(writer, sheet_name=name, startrow=gap_row_5th, startcol=0)

    # Customize fifth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_5th + 1,
        title="2024",
        last_row_color="D8B7DD"
    )

    # Add a gap of 4 rows before the sixth table
    gap_row_6th = gap_row_5th + r.shape[0] + 5
    # Save sixth DataFrame (accroissement_elec)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        accroissement_elec.to_excel(writer, sheet_name=name, startrow=gap_row_6th, startcol=0)

    # Customize sixth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_6th + 1,
        title="2024",
        last_row_color="90EE90"
    )

    # Add a gap of 4 rows before the seventh table
    gap_row_7th = gap_row_6th + accroissement_elec.shape[0] + 5
    # Save seventh DataFrame (tx_encaissement)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        ap.to_excel(writer, sheet_name=name, startrow=gap_row_7th, startcol=0)

    # Customize seventh DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_7th + 1,
        title="2024",
        last_row_color="FFFFE0"
    )



    # Add a gap of 4 rows before the eighth table
    gap_row_8th = gap_row_7th + ap.shape[0] + 5
    # Save eighth DataFrame (rea)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        tx_encaissement.to_excel(writer, sheet_name=name, startrow=gap_row_8th, startcol=0)

    # Customize eighth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_8th + 1,
        title="2024",
        last_row_color="FFA500"
    )

    # Add a gap of 4 rows before the ninth table
    gap_row_9th = gap_row_8th + tx_encaissement.shape[0] + 5
    # Save ninth DataFrame (apport_elec)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        rea.to_excel(writer, sheet_name=name, startrow=gap_row_9th, startcol=0)

    # Customize ninth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_9th + 1,
        title="2024",
        last_row_color="90EE90"
    )

    # Add a gap of 4 rows before the tenth table


    gap_row_10th = gap_row_9th + rea.shape[0] + 5
    # Save tenth DataFrame (apnv)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        apport_elec.to_excel(writer, sheet_name=name, startrow=gap_row_10th, startcol=0)

    # Customize tenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_10th + 1,
        title="2024",
        last_row_color="D3D3D3"
    )

    # Add a gap of 4 rows before the eleventh table
    gap_row_11th = gap_row_10th + apport_elec.shape[0] + 5
    # Save eleventh DataFrame (apport_gaz)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        apnv.to_excel(writer, sheet_name=name, startrow=gap_row_11th, startcol=0)

    # Customize eleventh DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_11th + 1,
        title="2024",
        last_row_color="FFFFE0"
    )


    # Add a gap of 4 rows before the twelfth table
    gap_row_12th = gap_row_11th + apnv.shape[0] + 5
    # Save twelfth DataFrame (ventesTB)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        apport_gaz.to_excel(writer, sheet_name=name, startrow=gap_row_12th, startcol=0)

    # Customize twelfth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_12th + 1,
        title="2024",
        last_row_color="90EE90"
    )

    # Add a gap of 4 rows before the thirteenth table
    gap_row_13th = gap_row_12th + apport_gaz.shape[0] + 5
    # Save thirteenth DataFrame (evolution)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        ventesTB.to_excel(writer, sheet_name=name, startrow=gap_row_13th, startcol=0)

    # Customize thirteenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_13th + 1,
        title="Evolution 2024",
        last_row_color="D8B7DD"
    )

    # Add a gap of 4 rows before the fourteenth table

    gap_row_14th = gap_row_13th + ventesTB.shape[0] + 5
    # Save fourteenth DataFrame (evolution2)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        evolution.to_excel(writer, sheet_name=name, startrow=gap_row_14th, startcol=0)

    # Customize fourteenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_14th + 1,
        title="Evolution 2024",
        last_row_color="FFC0CB"
    )


    # Add a gap of 4 rows before the fifteenth table
    gap_row_15th = gap_row_14th + evolution.shape[0] + 5
    # Save fifteenth DataFrame (achatsTB)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        achatsTB.to_excel(writer, sheet_name=name, startrow=gap_row_15th, startcol=0)

    # Customize fifteenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_15th + 1,
        title="2024",
        last_row_color="FFFFE0"
    )


    # Add a gap of 4 rows before the sixteenth table
    gap_row_16th = gap_row_15th + achatsTB.shape[0] + 5
    # Save sixteenth DataFrame (chiffaf)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        evolution2.to_excel(writer, sheet_name=name, startrow=gap_row_16th, startcol=0)

    # Customize sixteenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_16th + 1,
        title="2024",
        last_row_color="D3D3D3"
    )


    # Add a gap of 4 rows before the seventeenth table
    gap_row_17th = gap_row_16th + evolution2.shape[0] + 5
    # Save seventeenth DataFrame (prix)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        chiffaf.to_excel(writer, sheet_name=name, startrow=gap_row_17th, startcol=0)

    # Customize seventeenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_17th + 1,
        title="2024",
        last_row_color="F0E68C"
    )

 # Add a gap of 4 rows before the seventeenth table
    gap_row_18th = gap_row_17th + chiffaf.shape[0] + 5
    # Save seventeenth DataFrame (prix)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        chiffre_aff.to_excel(writer, sheet_name=name, startrow=gap_row_18th, startcol=0)

    # Customize seventeenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_18th + 1,
        title="2024",
        last_row_color="F0E68C"
    )

    # Add a gap of 4 rows before the seventeenth table
    gap_row_19th = gap_row_18th + chiffre_aff.shape[0] + 5
    # Save seventeenth DataFrame (prix)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        pr.to_excel(writer, sheet_name=name, startrow=gap_row_19th, startcol=0)

    # Customize seventeenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_19th + 1,
        title="2024",
        last_row_color="F0E68C"
    )


    # Add a gap of 4 rows before the seventeenth table
    gap_row_20th = gap_row_19th + pr.shape[0] + 5
    # Save seventeenth DataFrame (prix)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        prix.to_excel(writer, sheet_name=name, startrow=gap_row_20th, startcol=0)

    # Customize seventeenth DataFrame
    customize_excel_table(
        file_name=file_name,
        sheet_name=name,
        start_row=gap_row_20th + 1,
        title="2024",
        last_row_color="F0E68C"
    )

